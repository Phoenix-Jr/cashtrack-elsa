"""
Report generation utilities for PDF and XLSX formats
"""
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Sum, Q
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Transaction


def hex_to_rgb(hex_color):
    """Convert hex color to RGB tuple"""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))


def generate_pdf_report(transactions, report_type, period, start_date, end_date, user_name):
    """Generate a PDF report with beautiful design"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    
    # Define colors matching the app theme
    primary_blue = colors.HexColor('#0B74FF')
    primary_green = colors.HexColor('#10B981')
    primary_red = colors.HexColor('#EF4444')
    primary_purple = colors.HexColor('#8B5CF6')
    primary_orange = colors.HexColor('#F59E0B')
    dark_gray = colors.HexColor('#0F172A')
    light_gray = colors.HexColor('#F3F4F6')
    border_gray = colors.HexColor('#E5E7EB')
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=dark_gray,
        spaceAfter=12,
        alignment=1,  # Center
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=primary_blue,
        spaceAfter=8,
    )
    
    # Header with gradient effect
    header_data = [
        [Paragraph(f"<b>RAPPORT FINANCIER</b>", title_style)],
        [Paragraph(f"<i>{period}</i>", styles['Normal'])],
    ]
    header_table = Table(header_data, colWidths=[7*inch])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), primary_blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 20),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, 1), light_gray),
        ('FONTSIZE', (0, 1), (-1, 1), 12),
        ('TOPPADDING', (0, 1), (-1, 1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 8),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Calculate statistics
    total_recettes = sum(t.amount for t in transactions if t.type == 'recette')
    total_depenses = abs(sum(t.amount for t in transactions if t.type == 'depense'))
    balance = total_recettes - total_depenses
    transaction_count = len(transactions)
    
    # Summary Statistics
    summary_data = [
        ['Statistiques', 'Valeur'],
        ['Période', period],
        ['Date de début', start_date.strftime('%d/%m/%Y')],
        ['Date de fin', end_date.strftime('%d/%m/%Y')],
        ['Nombre de transactions', str(transaction_count)],
        ['', ''],
        ['Total Encaissements', f"{total_recettes:,.0f} FCFA"],
        ['Total Décaissements', f"{total_depenses:,.0f} FCFA"],
        ['Solde Net', f"{balance:,.0f} FCFA"],
    ]
    
    summary_table = Table(summary_data, colWidths=[3.5*inch, 3.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), primary_blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, border_gray),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light_gray]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('FONTNAME', (0, 6), (1, 6), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 6), (1, 6), primary_green),
        ('FONTNAME', (0, 7), (1, 7), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 7), (1, 7), primary_red),
        ('FONTNAME', (0, 8), (1, 8), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 8), (1, 8), primary_blue if balance >= 0 else primary_red),
    ]))
    story.append(Paragraph("<b>Résumé</b>", heading_style))
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Category Breakdown - Recettes
    recettes_by_category = {}
    for t in transactions:
        if t.type == 'recette':
            cat_name = t.category.name if t.category else 'Non catégorisé'
            recettes_by_category[cat_name] = recettes_by_category.get(cat_name, 0) + t.amount
    
    if recettes_by_category:
        story.append(Paragraph("<b>Répartition par Catégorie - Encaissements</b>", heading_style))
        cat_data = [['Catégorie', 'Montant']]
        for cat, amount in sorted(recettes_by_category.items(), key=lambda x: x[1], reverse=True):
            cat_data.append([cat, f"{amount:,.0f} FCFA"])
        
        cat_table = Table(cat_data, colWidths=[3.5*inch, 3.5*inch])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), primary_green),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, border_gray),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light_gray]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(cat_table)
        story.append(Spacer(1, 0.2*inch))
    
    # Category Breakdown - Dépenses
    depenses_by_category = {}
    for t in transactions:
        if t.type == 'depense':
            cat_name = t.category.name if t.category else 'Non catégorisé'
            depenses_by_category[cat_name] = depenses_by_category.get(cat_name, 0) + abs(t.amount)
    
    if depenses_by_category:
        story.append(Paragraph("<b>Répartition par Catégorie - Décaissements</b>", heading_style))
        cat_data = [['Catégorie', 'Montant']]
        for cat, amount in sorted(depenses_by_category.items(), key=lambda x: x[1], reverse=True):
            cat_data.append([cat, f"{amount:,.0f} FCFA"])
        
        cat_table = Table(cat_data, colWidths=[3.5*inch, 3.5*inch])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), primary_red),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, border_gray),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light_gray]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(cat_table)
        story.append(Spacer(1, 0.3*inch))
    
    # Transactions Table
    if transactions:
        story.append(Paragraph("<b>Détail des Transactions</b>", heading_style))
        trans_data = [['Date', 'Type', 'Description', 'Catégorie', 'Montant', 'Référence']]
        
        for t in transactions[:100]:  # Limit to 100 transactions per page
            trans_data.append([
                t.created_at.strftime('%d/%m/%Y'),
                'Entrée' if t.type == 'recette' else 'Sortie',
                t.description or 'N/A',
                t.category.name if t.category else 'Non catégorisé',
                f"{t.amount:,.0f} FCFA",
                t.ref,
            ])
        
        trans_table = Table(trans_data, colWidths=[1*inch, 0.8*inch, 2*inch, 1.2*inch, 1*inch, 1*inch])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), dark_gray),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, border_gray),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light_gray]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(trans_table)
    
    # Footer
    story.append(Spacer(1, 0.3*inch))
    footer_data = [
        [Paragraph(f"<i>Généré par: {user_name}</i>", styles['Normal'])],
        [Paragraph(f"<i>Date de génération: {datetime.now().strftime('%d/%m/%Y %H:%M')}</i>", styles['Normal'])],
    ]
    footer_table = Table(footer_data, colWidths=[7*inch])
    footer_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.grey),
    ]))
    story.append(footer_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_xlsx_report(transactions, report_type, period, start_date, end_date, user_name):
    """Generate an XLSX report with the same format as frontend"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport"
    
    # Define colors (matching frontend)
    primary_blue = "0B177C"  # Main blue from frontend
    light_blue = "F0F4FF"  # Light blue background
    gray = "B8B8B8"  # Gray borders
    light_gray = "F5F5F5"  # Alternating rows
    white = "FFFFFF"
    
    # Calculate statistics
    total_recettes = sum(t.amount for t in transactions if t.type == 'recette')
    total_depenses = abs(sum(t.amount for t in transactions if t.type == 'depense'))
    balance = total_recettes - total_depenses
    transaction_count = len(transactions)
    
    # Title row (row 1)
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = f"RAPPORT {report_type.upper()} - {period}".upper()
    title_cell.font = Font(bold=True, size=20, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=primary_blue, end_color=primary_blue, fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 45
    
    # Stats row (row 2)
    ws.merge_cells('A2:J2')
    stats_cell = ws['A2']
    stats_text = f"📊 Période: {period} | Total: {transaction_count} enregistrement(s) | Date d'export: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    stats_cell.value = stats_text
    stats_cell.font = Font(bold=True, size=14, color=primary_blue)
    stats_cell.fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
    stats_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 30
    
    # Column headers (row 3) - Same as frontend
    headers = [
        'ID', 'DATE DE CRÉATION', 'TYPE', 'CATÉGORIE', 'DESCRIPTION',
        'MONTANT', 'RÉFÉRENCE', 'EXPORTATEUR/FOURNISSEUR', 'SOLDE', 'CRÉÉ PAR'
    ]
    header_row = 3
    ws.row_dimensions[header_row].height = 40
    
    # Column widths (matching frontend)
    column_widths = [10, 20, 12, 18, 30, 15, 15, 25, 15, 18]
    
    for col_idx, (header, width) in enumerate(zip(headers, column_widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        # Handle EXPORTATEUR/FOURNISSEUR with line break
        if 'EXPORTATEUR/FOURNISSEUR' in header:
            cell.value = 'EXPORTATEUR/FOURNIS\nSEUR'
        else:
            cell.value = header
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color=primary_blue, end_color=primary_blue, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Add borders
        thin_border = Border(
            left=Side(style='medium', color=gray),
            right=Side(style='medium', color=gray),
            top=Side(style='thick', color=gray),
            bottom=Side(style='thick', color=gray)
        )
        cell.border = thin_border
    
    # Data rows (starting from row 4)
    data_start_row = 4
    current_balance = 0
    
    for row_idx, t in enumerate(transactions, start=data_start_row):
        ws.row_dimensions[row_idx].height = 25
        
        # Calculate balance
        if t.type == 'recette':
            current_balance += t.amount
        else:
            current_balance -= abs(t.amount)
        
        # Fill data
        data = [
            t.id,
            t.created_at.strftime('%d/%m/%Y %H:%M') if t.created_at else '',
            'Entrée' if t.type == 'recette' else 'Sortie',
            t.category.name if t.category else 'Non catégorisé',
            t.description or '',
            t.amount,
            t.ref or '',
            t.exporter_fournisseur or '',
            current_balance,
            t.created_by.name if t.created_by and hasattr(t.created_by, 'name') else (t.created_by.email if t.created_by else ''),
        ]
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx == 6 or col_idx == 9:  # MONTANT and SOLDE columns
                cell.value = float(value)
                cell.number_format = '#,##0" XOF"'
            else:
                cell.value = value
            
            cell.font = Font(size=11, color="000000")
            # Alternating row colors
            fill_color = white if row_idx % 2 == 0 else light_gray
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(
                horizontal='left' if col_idx == 1 else 'center',
                vertical='middle',
                wrap_text=True
            )
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin', color=gray),
                right=Side(style='thin', color=gray),
                top=Side(style='thin', color=gray),
                bottom=Side(style='thin', color=gray)
            )
            cell.border = thin_border
    
    # Totals row
    if transactions:
        totals_row = data_start_row + len(transactions)
        ws.row_dimensions[totals_row].height = 30
        
        totals_data = ['TOTAL', '', '', '', '', balance, '', '', balance, '']
        for col_idx, value in enumerate(totals_data, start=1):
            cell = ws.cell(row=totals_row, column=col_idx)
            if col_idx == 6 or col_idx == 9:  # MONTANT and SOLDE
                cell.value = float(value) if value else ''
                if value:
                    cell.number_format = '#,##0" XOF"'
            else:
                cell.value = value
            
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color=primary_blue, end_color=primary_blue, fill_type="solid")
            cell.alignment = Alignment(
                horizontal='left' if col_idx == 1 else 'center',
                vertical='middle',
                wrap_text=True
            )
            
            # Add borders
            thick_border = Border(
                left=Side(style='medium', color=gray),
                right=Side(style='medium', color=gray),
                top=Side(style='thick', color=gray),
                bottom=Side(style='thick', color=gray)
            )
            cell.border = thick_border
    
    # Statistics section
    stats_start_row = totals_row + 2 if transactions else data_start_row + len(transactions) + 1
    
    # Stats title
    ws.merge_cells(f'A{stats_start_row}:B{stats_start_row}')
    stats_title_cell = ws[f'A{stats_start_row}']
    stats_title_cell.value = '📊 STATISTIQUES'
    stats_title_cell.font = Font(bold=True, size=18, color="FFFFFF")
    stats_title_cell.fill = PatternFill(start_color=primary_blue, end_color=primary_blue, fill_type="solid")
    stats_title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[stats_start_row].height = 35
    
    thick_border = Border(
        left=Side(style='thick', color=gray),
        right=Side(style='thick', color=gray),
        top=Side(style='thick', color=gray),
        bottom=Side(style='thick', color=gray)
    )
    stats_title_cell.border = thick_border
    
    # Stats data
    stats_data = [
        ('Total opérations', transaction_count),
        ('Total recettes', f'{total_recettes:,.0f} XOF'),
        ('Total dépenses', f'{total_depenses:,.0f} XOF'),
        ('Solde', f'{balance:,.0f} XOF'),
    ]
    
    for idx, (label, value) in enumerate(stats_data, start=1):
        row = stats_start_row + idx
        ws.row_dimensions[row].height = 25
        
        # Label cell
        label_cell = ws.cell(row=row, column=1)
        label_cell.value = label
        label_cell.font = Font(bold=True, size=11, color=primary_blue)
        label_cell.alignment = Alignment(horizontal='left', vertical='middle')
        fill_color = white if idx % 2 == 0 else light_gray
        label_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        label_cell.border = Border(
            left=Side(style='thin', color=gray),
            right=Side(style='thin', color=gray),
            top=Side(style='thin', color=gray),
            bottom=Side(style='thin', color=gray)
        )
        
        # Value cell
        value_cell = ws.cell(row=row, column=2)
        value_cell.value = value
        value_cell.font = Font(bold=True, size=11, color=primary_blue)
        value_cell.alignment = Alignment(horizontal='left', vertical='middle')
        value_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        value_cell.border = Border(
            left=Side(style='thin', color=gray),
            right=Side(style='thin', color=gray),
            top=Side(style='thin', color=gray),
            bottom=Side(style='thin', color=gray)
        )
    
    # Adjust column widths for stats
    ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width or 0, 30)
    ws.column_dimensions['B'].width = max(ws.column_dimensions['B'].width or 0, 20)
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

